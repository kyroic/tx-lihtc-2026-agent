#!/usr/bin/env python3
"""
╔═══════════════════════════════════════════════════════════════════╗
║  LIHTC TX 2026 — Parcell Validation Pipeline                    ║
║  Single-file, idempotent, deterministic.                        ║
║  Usage: python3 pipeline.py                                     ║
╚═══════════════════════════════════════════════════════════════════╝

Data Sources:
  A. phase1_sheet_data.tsv         — 114 apps, ground truth (CT, lat/lng, amenities)
  B. parcell_9pct_data.json        — 86 scraped P9 entries (86 good, some error/timeout)
  C. parcell_4pct_data.json        — 105 scraped P4 entries (97 good, 8 errors)
  D. _search_results.json          — 65 address-based P4 searches (45 found, 42 CT match)
  E. _search_queue_fixed_results.json — 20 re-searched with simplified addresses (3 found, 2 CT match)

Merge Strategy:
  1. Collect ALL candidate data for each app across sources
  2. Pick BEST candidate: CT match > any data with CT > data without CT > nothing
  3. For ties: prefer the source with richer amenity/score data (P9 > P4_SEARCH > P4_DIRECT)

Output:
  Merged_Parcell_Data.json           — All 114 apps, unified schema
  Parcell_Validation_Workbook_pipeline.xlsx  — 14 sheets: 0_Summary, Master List, C1-C10, Not On Parcell, Coordinates
  Match_Validation_Workbook_pipeline.xlsx     — 8 sheets: Summary, Q, CT, Park, School, Grocery, Library, CT Verified
"""

import json, csv, re, os, sys, runpy, shutil
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════
# 1. LOAD ALL SOURCES
# ═══════════════════════════════════════════════════════════════════

def load_phase1(path='phase1_sheet_data.tsv'):
    """Load ground-truth Phase 1 data. Returns dict[app_name]."""
    data = {}
    with open(path, encoding='utf-8') as f:
        for row in csv.DictReader(f, delimiter='\t'):
            data[row['application_name']] = row
    return data

def load_p9(path='parcell_9pct_data.json'):
    """Load Parcell 9% scrape data. Returns dict[app_name]."""
    with open(path) as f:
        return json.load(f)

def load_p4(path='parcell_4pct_data.json'):
    """Load Parcell 4% scrape data. Returns dict[app_name]."""
    with open(path) as f:
        return json.load(f)

def load_search_results(path='_search_results.json'):
    """Load batch address-search results. Returns dict[app_name]."""
    with open(path) as f:
        return json.load(f)

def load_fixed_search(path='_search_queue_fixed_results.json'):
    """Load simplified-address search results. Returns dict[app_name]."""
    if not os.path.exists(path):
        return {}
    with open(path) as f:
        return json.load(f)


# ═══════════════════════════════════════════════════════════════════
# 2. CANDIDATE COLLECTION
# ═══════════════════════════════════════════════════════════════════

def collect_candidates(app_name, p1_row, p9, p4, search, fixed):
    """
    Collect all viable data candidates for a single app.
    
    Returns list of candidate dicts, each with:
      source, round, ct_match, parcel_census_tract, [source-specific fields]
    """
    candidates = []
    p1_ct = p1_row.get('census_tract', '').strip().replace('.0', '')
    
    # ── Candidate A: P9 direct scrape ──
    if app_name in p9:
        d = p9[app_name]
        if not d.get('error'):
            ct = str(d.get('census_tract', '')).strip()
            candidates.append({
                'source': 'P9_DIRECT',
                'round': '9%',
                'parcel_census_tract': ct,
                'ct_match': (p1_ct == ct) if p1_ct and ct else None,
                # P9-specific fields
                'quartile': d.get('quartile', ''),
                'poverty_rate': d.get('poverty_rate', ''),
                'poverty_rate_rank': d.get('poverty_rate_rank', ''),
                'total_score': d.get('total_score', ''),
                'park': d.get('park', ''),
                'school': d.get('school', ''),
                'grocery': d.get('grocery', ''),
                'library': d.get('library', ''),
                'amenity_tiebreaker': d.get('tie_breaker', ''),
                # Richness score for tiebreaking
                '_richness': 10,
            })
    
    # ── Candidate B: P4 direct scrape (valid PIDs only) ──
    if app_name in p4:
        d = p4[app_name]
        bad_pids = {'4', '960397', 'R4557', '0260470000106'}
        if not d.get('error') and str(d.get('pid', '')) not in bad_pids:
            ct = str(d.get('census_tract', '')).strip()
            candidates.append({
                'source': 'P4_DIRECT',
                'round': '4%',
                'parcel_id': str(d.get('pid', '')),
                'parcel_census_tract': ct,
                'ct_match': (p1_ct == ct) if p1_ct and ct else None,
                # P4-specific
                'parcel_address': d.get('address', ''),
                'parcel_city': '',
                'parcel_score': d.get('total_score', ''),
                'parcel_size': d.get('acreage', ''),
                'parcel_zoning': d.get('zoning', ''),
                'amenity_tiebreaker': d.get('tie_breaker', ''),
                'county': d.get('county', ''),
                'owner': d.get('owner', ''),
                'dda': d.get('dda', ''),
                'url': d.get('url', ''),
                '_richness': 5,
            })
    
    # ── Candidate C: P4 batch search (primary) ──
    if app_name in search:
        d = search[app_name]
        if d.get('found'):
            ct = str(d.get('census_tract', '')).strip()
            candidates.append({
                'source': 'P4_SEARCH',
                'round': '4%',
                'parcel_id': str(d.get('parcel_id', '')),
                'parcell_uid': d.get('parcell_uid', ''),
                'parcel_census_tract': ct,
                'ct_match': d.get('ct_match', False),
                'parcel_address': d.get('parcel_address', ''),
                'parcel_city': d.get('parcel_city', ''),
                'parcel_score': d.get('score', ''),
                'parcel_size': d.get('size', ''),
                'parcel_zoning': d.get('zoning', ''),
                'query_address': d.get('address', ''),
                '_richness': 7,
            })
    
    # ── Candidate D: P4 fixed-address search ──
    if app_name in fixed:
        d = fixed[app_name]
        if d.get('found'):
            ct = str(d.get('census_tract', '')).strip()
            candidates.append({
                'source': 'P4_SEARCH_FIXED',
                'round': '4%',
                'parcel_id': str(d.get('parcel_id', '')),
                'parcell_uid': d.get('parcell_uid', ''),
                'parcel_census_tract': ct,
                'ct_match': d.get('ct_match', False),
                'parcel_address': d.get('parcel_address', ''),
                'parcel_city': d.get('parcel_city', ''),
                'parcel_score': d.get('score', ''),
                'parcel_size': d.get('size', ''),
                'parcel_zoning': d.get('zoning', ''),
                'query_address': d.get('address', ''),
                '_richness': 6,
            })
    
    return candidates


def select_best_candidate(candidates):
    """
    Select the best candidate from all sources.
    
    Priority:
      1. Any CT match (prefer highest richness)
      2. Any data with a CT to compare (prefer highest richness)
      3. Any data without CT (fallback)
      4. Nothing
    """
    if not candidates:
        return None
    
    def sort_key(c):
        # Primary: CT match (True > None > False)
        match_rank = 0 if c['ct_match'] is True else (1 if c['ct_match'] is None else 2)
        # Secondary: richness (higher is better)
        richness = -c.get('_richness', 0)
        return (match_rank, richness)
    
    candidates.sort(key=sort_key)
    return candidates[0]


# ═══════════════════════════════════════════════════════════════════
# 3. MERGE & CLASSIFY
# ═══════════════════════════════════════════════════════════════════

def merge_all(p1, p9, p4, search, fixed):
    """Merge all sources into a single dict of 114 apps."""
    merged = {}
    
    for app_name in sorted(p1.keys()):
        p1_row = p1[app_name]
        p1_ct = p1_row.get('census_tract', '').strip().replace('.0', '')
        
        # Base entry
        entry = {
            'app_name': app_name,
            'p1_ct': p1_ct,
            'p1_lat': p1_row.get('site_lat', ''),
            'p1_lng': p1_row.get('site_lng', ''),
            'p1_quartile': p1_row.get('quartile', ''),
            'p1_tiebreaker': p1_row.get('tiebreaker_score', ''),
            'p1_poverty_rank': p1_row.get('poverty_rank', ''),
        }
        
        candidates = collect_candidates(app_name, p1_row, p9, p4, search, fixed)
        best = select_best_candidate(candidates)
        
        if best:
            # Copy all fields from best candidate
            for k, v in best.items():
                if not k.startswith('_'):
                    entry[k] = v
            
            # Assign status
            if best['ct_match'] is True:
                entry['status'] = f"VERIFIED_{best['source']}"
            elif best['ct_match'] is False:
                entry['status'] = f"{best['source']}_CT_MISMATCH"
            else:
                entry['status'] = f"FALLBACK_{best['source']}"
        else:
            entry.update({
                'source': 'NONE',
                'round': '',
                'parcel_census_tract': '',
                'ct_match': None,
                'status': 'NO_DATA',
            })
        
        merged[app_name] = entry
    
    return merged


# ═══════════════════════════════════════════════════════════════════
# 4. STATISTICS
# ═══════════════════════════════════════════════════════════════════

def compute_stats(merged):
    """Compute summary statistics."""
    total = len(merged)
    
    verified = [v for v in merged.values() if v['status'].startswith('VERIFIED')]
    mismatch = [v for v in merged.values() if 'MISMATCH' in v['status']]
    fallback = [v for v in merged.values() if v['status'].startswith('FALLBACK')]
    no_data = [v for v in merged.values() if v['status'] == 'NO_DATA']
    
    by_source = defaultdict(lambda: {'total': 0, 'ct_match': 0, 'ct_mismatch': 0, 'ct_none': 0})
    for v in merged.values():
        src = v.get('source', 'NONE')
        by_source[src]['total'] += 1
        if v.get('ct_match') is True:
            by_source[src]['ct_match'] += 1
        elif v.get('ct_match') is False:
            by_source[src]['ct_mismatch'] += 1
        else:
            by_source[src]['ct_none'] += 1
    
    return {
        'total': total,
        'verified': len(verified),
        'verified_p9': len([v for v in verified if v['source'] == 'P9_DIRECT']),
        'verified_p4_direct': len([v for v in verified if v['source'] == 'P4_DIRECT']),
        'verified_p4_search': len([v for v in verified if v['source'] == 'P4_SEARCH']),
        'verified_p4_fixed': len([v for v in verified if v['source'] == 'P4_SEARCH_FIXED']),
        'mismatch': len(mismatch),
        'fallback': len(fallback),
        'no_data': len(no_data),
        'by_source': dict(by_source),
    }


# ═══════════════════════════════════════════════════════════════════
# 5. EXCEL WORKBOOK
# ═══════════════════════════════════════════════════════════════════

# Style constants
HDR_FILL = PatternFill('solid', fgColor='4472C4')
HDR_FONT = Font(bold=True, size=11, color='FFFFFF')
GREEN = PatternFill('solid', fgColor='C6EFCE')
YELLOW = PatternFill('solid', fgColor='FFEB9C')
RED = PatternFill('solid', fgColor='FFC7CE')
GRAY = PatternFill('solid', fgColor='D9D9D9')
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(wrap_text=True, horizontal='center', vertical='top')


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = THIN


def auto_width(ws, max_w=55):
    for col_cells in ws.columns:
        mx = 0
        for cell in col_cells:
            if cell.value:
                mx = max(mx, min(max_w, len(str(cell.value))))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = mx + 3


def status_fill(status):
    if 'VERIFIED' in status: return GREEN
    if 'FALLBACK' in status: return YELLOW
    if 'MISMATCH' in status: return RED
    return GRAY


def build_validation_workbook(merged, stats, output_path):
    """Build pipeline 14-tab Parcell workbook via refactored 13-tab+Coordinates builder."""
    generated = 'Parcell_Validation_Workbook_pipeline.xlsx'
    runpy.run_path('_build_pipeline_13tab.py', run_name='__main__')
    if output_path != generated:
        shutil.copyfile(generated, output_path)
    return output_path


def build_match_workbook(merged, output_path):
    """Build pipeline multi-tab Match workbook via refactored per-field tab builder."""
    generated = 'Match_Validation_Workbook_pipeline.xlsx'
    runpy.run_path('_build_match_multi.py', run_name='__main__')
    if output_path != generated:
        shutil.copyfile(generated, output_path)
    return output_path


# ═══════════════════════════════════════════════════════════════════
# 6. MAIN
# ═══════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("LIHTC TX 2026 — PARCELL VALIDATION PIPELINE")
    print("=" * 60)
    
    # Load
    print("\n[1/5] Loading sources...")
    p1 = load_phase1()
    p9 = load_p9()
    p4 = load_p4()
    search = load_search_results()
    fixed = load_fixed_search()
    print(f"  Phase 1: {len(p1)} apps")
    print(f"  P9 data: {len(p9)} entries ({len([v for v in p9.values() if not v.get('error')])} good, {len([v for v in p9.values() if v.get('error')])} errors)")
    print(f"  P4 data: {len(p4)} entries ({len([v for v in p4.values() if not v.get('error')])} good, {len([v for v in p4.values() if v.get('error')])} errors)")
    print(f"  Batch search: {len(search)} searched ({sum(1 for v in search.values() if v.get('found'))} found)")
    print(f"  Fixed search: {len(fixed)} searched ({sum(1 for v in fixed.values() if v.get('found'))} found)")
    
    # Merge
    print("\n[2/5] Merging candidates & selecting best...")
    merged = merge_all(p1, p9, p4, search, fixed)
    assert len(merged) == len(p1), f"Expected {len(p1)} apps, got {len(merged)}"
    
    # Stats
    print("\n[3/5] Computing statistics...")
    stats = compute_stats(merged)
    
    print(f"\n  ┌─────────────────────────────────────┐")
    print(f"  │  VERIFIED (CT match): {stats['verified']:3d} / {stats['total']:3d}  ({stats['verified']/stats['total']*100:.1f}%)  │")
    print(f"  │    P9 Direct:             {stats['verified_p9']:3d}                 │")
    print(f"  │    P4 Direct:             {stats['verified_p4_direct']:3d}                 │")
    print(f"  │    P4 Search:             {stats['verified_p4_search']:3d}                 │")
    print(f"  │    P4 Fixed Search:       {stats['verified_p4_fixed']:3d}                 │")
    print(f"  ├─────────────────────────────────────┤")
    print(f"  │  CT MISMATCH:           {stats['mismatch']:3d}                      │")
    print(f"  │  FALLBACK:               {stats['fallback']:3d}                      │")
    print(f"  │  NO DATA:                {stats['no_data']:3d}                      │")
    print(f"  └─────────────────────────────────────┘")
    
    # Mismatch details
    mismatches = [(k, v) for k, v in merged.items() if 'MISMATCH' in v['status']]
    if mismatches:
        print("\n  CT Mismatches:")
        for name, v in sorted(mismatches):
            print(f"    {name}")
            print(f"      P1: {v['p1_ct']}  |  Parcell: {v['parcel_census_tract']}  |  Source: {v['source']}")
    
    no_data = [(k, v) for k, v in merged.items() if v['status'] == 'NO_DATA']
    if no_data:
        print(f"\n  No Data ({len(no_data)}):")
        for name, v in sorted(no_data):
            print(f"    {name}  (P1 CT: {v['p1_ct']}, lat: {v['p1_lat']}, lng: {v['p1_lng']})")
    
    # Save JSON
    print("\n[4/5] Saving Merged_Parcell_Data.json...")
    with open('Merged_Parcell_Data.json', 'w') as f:
        json.dump(merged, f, indent=2, default=str)
    print("  Done.")
    
    # Build workbooks
    print("\n[5/5] Building Excel workbooks...")
    build_validation_workbook(merged, stats, 'Parcell_Validation_Workbook_pipeline.xlsx')
    print("  ✅ Parcell_Validation_Workbook_pipeline.xlsx")
    build_match_workbook(merged, 'Match_Validation_Workbook_pipeline.xlsx')
    print("  ✅ Match_Validation_Workbook_pipeline.xlsx")
    
    print("\n" + "=" * 60)
    print("PIPELINE COMPLETE")
    print("=" * 60)
    return merged, stats


if __name__ == '__main__':
    main()
