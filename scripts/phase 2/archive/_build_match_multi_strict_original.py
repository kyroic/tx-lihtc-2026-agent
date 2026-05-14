#!/usr/bin/env python3
"""Build 8-tab Match_Validation_Workbook_RESTORED.xlsx using strict original on/off logic
(from _validation_pipeline.py style), yielding 86 on-Parcell + 28 off-Parcell apps.
"""
import json, csv, re
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('phase1_sheet_data.tsv', encoding='latin-1') as f:
    p1_rows = list(csv.DictReader(f, delimiter='\t'))
with open('parcell_9pct_data.json') as f:
    p9 = json.load(f)
with open('parcell_4pct_data.json') as f:
    p4 = json.load(f)

# styles
gf = PatternFill('solid', fgColor='C6EFCE')
rf = PatternFill('solid', fgColor='FFC7CE')
gr = PatternFill('solid', fgColor='D9D9D9')
bf = PatternFill('solid', fgColor='4472C4')
hf = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
nf = Font(name='Calibri', size=9)
bdr = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
ca = Alignment(horizontal='center', vertical='top', wrap_text=True)
qt_map = {'1st': '1', '2nd': '2', '3rd': '3', '4th': '4'}
amenities = ['park', 'school', 'grocery', 'library']


def is_blank(val):
    if not val or val == 'None':
        return True
    v2 = str(val).strip()
    if not v2 or v2.lower() == 'data unavailable':
        return True
    bare = re.sub(r'\s+', ' ', v2)
    return bare in ('0 ft', '0', '0ft', '0mi', '0.0 ft', '0.0', '0.0ft', '0.0mi')


def parse_ft(val):
    if not val:
        return None
    m = re.search(r'([\d,.]+)\s*ft', str(val))
    if m:
        return float(m.group(1).replace(',', ''))
    m = re.search(r'([\d,.]+)\s*mi', str(val))
    if m:
        return float(m.group(1).replace(',', '')) * 5280
    return None


# duplicate PID detection from original validation logic
pid_count = defaultdict(int)
for _name, d in p4.items():
    if d.get('error'):
        continue
    pid = str(d.get('pid', '')).strip()
    if pid and pid != 'None':
        pid_count[pid] += 1
dupe_pids = {pid for pid, c in pid_count.items() if c >= 2}


records = []
for p1d in p1_rows:
    name = p1d['application_name']
    p9d = p9.get(name, {})
    p4d = p4.get(name, {})

    p9_ok = bool(p9d and not p9d.get('error'))
    p4_ok = bool(p4d and not p4d.get('error'))
    p4_pid = str(p4d.get('pid', '')).strip() if p4d else ''

    p4_pid_is_trash = False
    if not p4_pid or p4_pid == 'None':
        p4_pid_is_trash = True
    elif any(c.isalpha() for c in p4_pid) or (p4_pid.isdigit() and len(p4_pid) < 5):
        p4_pid_is_trash = True
    elif p4_pid in dupe_pids:
        p4_pid_is_trash = True

    p4_has_ct = bool(p4_ok and not is_blank(p4d.get('census_tract', '')))
    if p4_pid in dupe_pids:
        p4_has_ct = False

    on_parcell = p9_ok or (p4_ok and (not p4_pid_is_trash or p4_has_ct))

    if p9_ok:
        psrc, pd = '9%', p9d
    elif p4_ok and on_parcell:
        psrc, pd = '4%', p4d
    else:
        psrc, pd = 'N/A', {}

    p1_ct = (p1d.get('census_tract', '') or '').strip()
    p_ct = str(pd.get('census_tract', '') or '').strip() if pd else ''

    if on_parcell and p1_ct and p_ct:
        ct_match = 'YES' if p1_ct == p_ct else 'NO'
    else:
        ct_match = 'N/A'

    q_match = 'N/A'
    if on_parcell and pd:
        p_qt = pd.get('quartile', '')
        p1_qt = p1d.get('quartile', '')
        if p_qt and p_qt != 'Data Unavailable' and p1_qt:
            p_qn = qt_map.get(p_qt, p_qt)
            q_match = 'YES' if p1_qt == p_qn else 'NO'

    am = {}
    for a in amenities:
        if on_parcell and pd:
            p1_v = (p1d.get(f'distance_to_{a}', '') or '').strip()
            p_ft = parse_ft(pd.get(a, ''))
            if p1_v and p_ft is not None and p_ft > 0:
                try:
                    p1_ft = float(p1_v)
                    ratio = p1_ft / p_ft
                    am[a] = 'YES' if 0.90 <= ratio <= 1.10 else 'NO'
                except Exception:
                    am[a] = 'N/A'
            else:
                am[a] = 'N/A'
        else:
            am[a] = 'N/A'

    if ct_match == 'YES':
        ct_label = '✅ CT MATCH'
    elif ct_match == 'NO':
        ct_label = '⚠️ CT MISMATCH'
    else:
        ct_label = '❌ NO RESULTS'

    records.append({
        'name': name,
        'on_parcell': on_parcell,
        'round': psrc,
        'q_match': q_match,
        'ct_match': ct_match,
        'ct_label': ct_label,
        'p1_ct': p1_ct,
        'p_ct': p_ct,
        'parcel_id': str(pd.get('pid', '')) if pd else '',
        'parcel_address': pd.get('address', '') if pd else '',
        'p1_quartile': p1d.get('quartile', ''),
        'p_quartile': pd.get('quartile', '') if pd else '',
        **{f'am_{a}': am[a] for a in amenities},
        **{f'p1_dist_{a}': p1d.get(f'distance_to_{a}', '') for a in amenities},
        **{f'p_dist_{a}': pd.get(a, '') if pd else '' for a in amenities},
    })

on_count = sum(1 for r in records if r['on_parcell'])
off_count = len(records) - on_count
ct_yes = sum(1 for r in records if r['ct_match'] == 'YES')
ct_no = sum(1 for r in records if r['ct_match'] == 'NO')
ct_na = sum(1 for r in records if r['ct_match'] == 'N/A')

match_fields = [
    ('Q', 'q_match', 'Quartile Match', 'Quartile ranking match between Phase 1 and Parcell'),
    ('CT', 'ct_match', 'CT Match', 'Census Tract match (direct data comparison)'),
    ('Park', 'am_park', 'Park Distance Match', 'Park distance match within ±10% tolerance'),
    ('School', 'am_school', 'School Distance Match', 'School distance match within ±10% tolerance'),
    ('Grocery', 'am_grocery', 'Grocery Distance Match', 'Grocery distance match within ±10% tolerance'),
    ('Library', 'am_library', 'Library Distance Match', 'Library distance match within ±10% tolerance'),
]

wb = Workbook()
ws = wb.active
ws.title = 'Summary'
ws.merge_cells('A1:F1')
ws['A1'] = 'LIHTC TX 2026 — Match Validation Summary (STRICT ORIGINAL)'
ws['A1'].font = Font(name='Calibri', size=16, bold=True)
ws.merge_cells('A2:F2')
ws['A2'] = f'{on_count} on + {off_count} off = {len(records)} total | Summary table below is ON-PARCELL ({on_count}) only'
ws['A2'].font = Font(name='Calibri', size=10, italic=True)

for col, h in enumerate(['Field', 'Description', 'YES (Match)', 'NO (Mismatch)', 'N/A', 'Match Rate'], 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = hf
    c.fill = bf
    c.border = bdr
    c.alignment = ca

row = 5
for label, key, _title, desc in match_fields:
    cnt = Counter(r[key] for r in records if r['on_parcell'])
    y = cnt.get('YES', 0)
    n = cnt.get('NO', 0)
    na = cnt.get('N/A', 0)
    rate = f'{(y/(y+n)*100):.1f}%' if (y+n) else 'N/A'
    vals = [label, desc, y, n, na, rate]
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = nf
        c.border = bdr
    row += 1

for j, w in enumerate([12, 55, 14, 14, 8, 12]):
    ws.column_dimensions[get_column_letter(j + 1)].width = w

for label, key, title, desc in match_fields:
    ws2 = wb.create_sheet(label)
    ws2.merge_cells('A1:F1')
    ws2['A1'] = f'{title} — Match Validation'
    ws2['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws2.merge_cells('A2:F2')
    ws2['A2'] = desc
    ws2['A2'].font = Font(name='Calibri', size=10, italic=True)

    for col, h in enumerate(['#', 'Name', 'Round', 'P1 Value', 'Parcell Value', 'Match'], 1):
        c = ws2.cell(row=4, column=col, value=h)
        c.font = hf
        c.fill = bf
        c.border = bdr
        c.alignment = ca

    rr = 5
    for i, r in enumerate(records):
        if label == 'Q':
            p1v, pv = r['p1_quartile'], r['p_quartile']
        elif label == 'CT':
            p1v, pv = r['p1_ct'], r['p_ct']
        else:
            a = label.lower()
            p1v, pv = r[f'p1_dist_{a}'], r[f'p_dist_{a}']

        vals = [i + 1, r['name'], r['round'], p1v, pv, r[key]]
        for col, v in enumerate(vals, 1):
            c = ws2.cell(row=rr, column=col, value=v)
            c.font = nf
            c.border = bdr
            if col == 6:
                c.alignment = ca
                c.fill = {'YES': gf, 'NO': rf}.get(r[key], gr)
        rr += 1

    for j, w in enumerate([5, 40, 8, 20, 40, 14]):
        ws2.column_dimensions[get_column_letter(j + 1)].width = w
    ws2.freeze_panes = 'A5'

ws_ct = wb.create_sheet('CT Verified')
ws_ct.merge_cells('A1:H1')
ws_ct['A1'] = 'CT Verification Results (STRICT ORIGINAL)'
ws_ct['A1'].font = Font(name='Calibri', size=14, bold=True)
ws_ct.merge_cells('A2:H2')
ws_ct['A2'] = f'{ct_yes}✅ CT MATCH | {ct_no}⚠️ CT MISMATCH | {ct_na}❌ NO RESULTS'
ws_ct['A2'].font = Font(name='Calibri', size=10, italic=True)

for col, h in enumerate(['#', 'Name', 'Round', 'P1 CT', 'Parcell CT', 'CT Verified', 'Parcell ID', 'Parcell Address'], 1):
    c = ws_ct.cell(row=4, column=col, value=h)
    c.font = hf
    c.fill = bf
    c.border = bdr
    c.alignment = ca

for i, r in enumerate(records):
    rw = i + 5
    vals = [i + 1, r['name'], r['round'], r['p1_ct'], r['p_ct'], r['ct_label'], r['parcel_id'], r['parcel_address']]
    for col, v in enumerate(vals, 1):
        c = ws_ct.cell(row=rw, column=col, value=v)
        c.font = nf
        c.border = bdr
        if col == 6:
            c.alignment = ca
            if 'MATCH' in str(v):
                c.fill = gf
            elif 'MISMATCH' in str(v):
                c.fill = rf
            else:
                c.fill = gr

for j, w in enumerate([5, 40, 8, 14, 14, 18, 14, 35]):
    ws_ct.column_dimensions[get_column_letter(j + 1)].width = w

out = 'Match_Validation_Workbook_RESTORED.xlsx'
wb.save(out)
print(f'Wrote: {out}')
print('Sheets:', wb.sheetnames)
print(f'on={on_count}, off={off_count}, CT yes={ct_yes}, no={ct_no}, na={ct_na}')
