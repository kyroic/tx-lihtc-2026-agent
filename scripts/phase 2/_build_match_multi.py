#!/usr/bin/env python3
"""Build multi-tab Match_Validation_Workbook_pipeline.xlsx matching original format
with one tab per comparison column: Q Match, CT Match, Park, School, Grocery, Library.
Uses pipeline CT-verified status as the primary match signal, plus amenity distance
comparison for park/school/grocery/library tabs."""
import json, csv, re, argparse
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

parser = argparse.ArgumentParser(description='Build multi-tab Match validation workbook')
parser.add_argument('--merged', default='Merged_Parcell_Data_pipeline_rerun.json',
                    help='Merged Parcell JSON source (pipeline rerun or original merge)')
parser.add_argument('--out', default='Match_Validation_Workbook_pipeline.xlsx',
                    help='Output workbook path')
args = parser.parse_args()

with open('phase1_sheet_data.tsv', encoding='latin-1') as f:
    p1_rows = list(csv.DictReader(f, delimiter='\t'))

with open(args.merged) as f: pr = json.load(f)
with open('parcell_9pct_data.json') as f: p9 = json.load(f)
with open('parcell_4pct_data.json') as f: p4 = json.load(f)

# Styles
gf = PatternFill('solid', fgColor='C6EFCE')
rf = PatternFill('solid', fgColor='FFC7CE')
yf = PatternFill('solid', fgColor='FFEB9C')
gr = PatternFill('solid', fgColor='D9D9D9')
bf = PatternFill('solid', fgColor='4472C4')
hf = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
nf = Font(name='Calibri', size=9)
bdr = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
ca = Alignment(horizontal='center', vertical='top', wrap_text=True)

qt_map = {'1st':'1','2nd':'2','3rd':'3','4th':'4'}
amenities = ['park','school','grocery','library']
amenity_labels = {'park':'Park','school':'School','grocery':'Grocery','library':'Library'}

def parse_ft(val):
    if not val: return None
    m = re.search(r'([\d,.]+)\s*ft', str(val))
    if m: return float(m.group(1).replace(',',''))
    m = re.search(r'([\d,.]+)\s*mi', str(val))
    if m: return float(m.group(1).replace(',','')) * 5280
    return None

def style_header(ws, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(1, c)
        cell.font = hf; cell.fill = bf; cell.border = bdr; cell.alignment = ca

def auto_width(ws, max_w=55):
    for col_cells in ws.columns:
        w2 = 0
        for cell in col_cells:
            if cell.value: w2 = max(w2, min(max_w, len(str(cell.value))))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = w2 + 3

# Build records list
records = []
for p1d in p1_rows:
    name = p1d['application_name']
    pr_entry = pr.get(name, {})
    p9d = p9.get(name, {})
    p4d = p4.get(name, {})
    
    p9_ok = bool(p9d and not p9d.get('error') and p9d.get('census_tract'))
    p4_ok = bool(p4d and not p4d.get('error') and p4d.get('census_tract'))
    
    if p9_ok: psrc, pd = '9%', p9d
    elif p4_ok: psrc, pd = '4%', p4d
    else: psrc, pd = '', {}
    
    # CT match from pipeline
    pipe_status = pr_entry.get('status', '')
    ct_verified = False
    if 'VERIFIED' in pipe_status: ct_verified = True
    ct_mismatch = 'MISMATCH' in pipe_status
    
    pipe_pid = pr_entry.get('parcel_id','') or pr_entry.get('parcell_id','')
    pipe_addr = pr_entry.get('parcel_address','') or pr_entry.get('parcell_address','')
    pipe_city = pr_entry.get('parcel_city','') or pr_entry.get('parcell_city','')
    pipe_score = str(pr_entry.get('parcel_score','') or pr_entry.get('parcell_score',''))
    
    # Quartile match
    q_match = 'N/A'
    if pd:
        p_qt = pd.get('quartile','')
        p1_qt = p1d.get('quartile','')
        if p_qt and p_qt != 'Data Unavailable' and p1_qt:
            p_qn = qt_map.get(p_qt, p_qt)
            q_match = 'YES' if p1_qt == p_qn else 'NO'
    
    # CT match
    if pd:
        p_ct = str(pd.get('census_tract','')).strip()
        p1_ct = (p1d.get('census_tract','') or '').strip()
        if p1_ct and p_ct:
            ct_match = 'YES' if p1_ct == p_ct else 'NO'
        else:
            ct_match = 'N/A'
    else:
        p1_ct = (p1d.get('census_tract','') or '').strip()
        ct_match = 'N/A'
    
    # Amenity matches
    am_matches = {}
    for a in amenities:
        if pd:
            p1_v = p1d.get(f'distance_to_{a}','').strip()
            p_ft = parse_ft(pd.get(a,''))
            if p1_v and p_ft is not None and p_ft > 0:
                try:
                    p1_ft = float(p1_v)
                    r = p1_ft / p_ft
                    am_matches[a] = 'YES' if 0.90 <= r <= 1.10 else 'NO'
                except: am_matches[a] = 'N/A'
            else:
                am_matches[a] = 'N/A'
        else:
            am_matches[a] = 'N/A'
    
    records.append({
        'name': name, 'round': psrc, 'pd': pd,
        'q_match': q_match, 'ct_match': ct_match,
        'ct_verified': ct_verified, 'ct_mismatch': ct_mismatch,
        'p1_ct': p1d.get('census_tract',''),
        'p_ct': str(pd.get('census_tract','')) if pd else '',
        'pipe_pid': pipe_pid, 'pipe_addr': pipe_addr, 'pipe_city': pipe_city,
        'pipe_score': pipe_score, 'pipe_ct_verified': ct_verified,
        **{f'am_{a}': am_matches[a] for a in amenities},
        'p1_quartile': p1d.get('quartile',''),
        'p_quartile': pd.get('quartile','') if pd else '',
        **{f'p1_dist_{a}': p1d.get(f'distance_to_{a}','') for a in amenities},
        **{f'p_dist_{a}': pd.get(a,'') if pd else '' for a in amenities},
    })

total = len(records)
match_fields = [
    ('Q', 'q_match', 'Quartile Match', 'Quartile ranking match between Phase 1 and Parcell'),
    ('CT', 'ct_match', 'CT Match', 'Census Tract match (direct data comparison)'),
    ('Park', 'am_park', 'Park Distance Match', 'Park distance match within ±10% tolerance'),
    ('School', 'am_school', 'School Distance Match', 'School distance match within ±10% tolerance'),
    ('Grocery', 'am_grocery', 'Grocery Distance Match', 'Grocery distance match within ±10% tolerance'),
    ('Library', 'am_library', 'Library Distance Match', 'Library distance match within ±10% tolerance'),
]

def match_fill(val):
    return {'YES': gf, 'NO': rf}.get(val, gr)

base_cols = ['#', 'Name', 'Round', 'P1 Value', 'Parcell Value', 'Match']

wb = Workbook()

# ═══════════ Summary ═══════════
ws = wb.active
ws.title = 'Summary'
ws.merge_cells('A1:F1')
ws['A1'] = 'LIHTC TX 2026 — Match Validation Summary'
ws['A1'].font = Font(name='Calibri', size=16, bold=True)
ws.merge_cells('A2:F2')
ct_yes = sum(1 for r in records if r['ct_verified'])
ct_no = sum(1 for r in records if r['ct_mismatch'])
ct_na = total - ct_yes - ct_no
source_tag = 'Pipeline CT' if 'pipeline' in args.merged.lower() else 'CT'
ws['A2'] = f'{total} applications | {source_tag}: {ct_yes}✅ Verified, {ct_no}⚠️ Mismatch, {ct_na}❌ No Results'
ws['A2'].font = Font(name='Calibri', size=10, italic=True)

for col, h in enumerate(['Field', 'Description', 'YES (Match)', 'NO (Mismatch)', 'N/A', 'Match Rate'], 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = hf; c.fill = bf; c.border = bdr; c.alignment = ca

row = 5
overall_yes = overall_no = 0
for label, key, title, desc in match_fields:
    counts = Counter(r[key] for r in records)
    yes_c = counts.get('YES', 0)
    no_c = counts.get('NO', 0)
    na_c = counts.get('N/A', 0)
    rate = f'{yes_c/(yes_c+no_c)*100:.1f}%' if yes_c+no_c else 'N/A'
    overall_yes += yes_c; overall_no += no_c
    data = [label, desc, yes_c, no_c, na_c, rate]
    for col, v in enumerate(data, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = nf; c.border = bdr
        if col in (1, 6): c.font = Font(name='Calibri', size=9, bold=True)
    row += 1

overall_rate = f'{overall_yes/(overall_yes+overall_no)*100:.1f}%' if overall_yes+overall_no else 'N/A'
row += 1
for col, v in enumerate(['OVERALL', '', overall_yes, overall_no, '', overall_rate], 1):
    c = ws.cell(row=row, column=col, value=v)
    c.font = Font(name='Calibri', size=10, bold=True); c.border = bdr
    if col in (3, 6): c.fill = gf

widths = [12, 55, 14, 14, 8, 12]
for j, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(j+1)].width = w

# ═══════════ Per-field tabs ═══════════
for label, key, title, desc in match_fields:
    ws2 = wb.create_sheet(label)
    ws2.merge_cells('A1:F1')
    ws2['A1'] = f'{title} — Match Validation'
    ws2['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws2.merge_cells('A2:F2')
    ws2['A2'] = desc
    ws2['A2'].font = Font(name='Calibri', size=10, italic=True)
    
    for col, h in enumerate(base_cols, 1):
        c = ws2.cell(row=4, column=col, value=h)
        c.font = hf; c.fill = bf; c.border = bdr; c.alignment = ca
    
    row2 = 5
    for i, r in enumerate(records):
        match_val = r[key]
        
        # Get P1 and Parcell values based on field type
        if label == 'Q':
            p1_val = r['p1_quartile']
            p_val = r['p_quartile']
        elif label == 'CT':
            p1_val = r['p1_ct']
            p_val = r['p_ct']
        elif label in ('Park', 'School', 'Grocery', 'Library'):
            al = label.lower()
            p1_val = r[f'p1_dist_{al}']
            p_val = str(r[f'p_dist_{al}'])
        else:
            p1_val = p_val = ''
        
        vals = [i+1, r['name'], r['round'], p1_val, p_val, match_val]
        for col, v in enumerate(vals, 1):
            c = ws2.cell(row=row2, column=col, value=v)
            c.font = nf; c.border = bdr
            if col == 6:
                c.fill = match_fill(match_val); c.alignment = ca
        row2 += 1
    
    # Summary at bottom
    counts = Counter(r[key] for r in records)
    row2 += 2
    for col, v in enumerate(['SUMMARY', f'YES={counts.get("YES",0)}', f'NO={counts.get("NO",0)}', f'N/A={counts.get("N/A",0)}', '', ''], 1):
        c = ws2.cell(row=row2, column=col, value=v)
        c.font = Font(name='Calibri', size=10, bold=True)
    
    widths2 = [5, 40, 8, 20, 40, 14]
    for j, w in enumerate(widths2):
        ws2.column_dimensions[get_column_letter(j+1)].width = w
    ws2.freeze_panes = 'A5'

# ═══════════ CT Verified (Pipeline) ═══════════
ws_ct = wb.create_sheet('CT Verified')
ws_ct.merge_cells('A1:H1')
ws_ct['A1'] = 'CT Verification Results'
ws_ct['A1'].font = Font(name='Calibri', size=14, bold=True)
ws_ct.merge_cells('A2:H2')
ws_ct['A2'] = f'{ct_yes}✅ CT MATCH | {ct_no}⚠️ CT MISMATCH | {ct_na}❌ NO RESULTS'
ws_ct['A2'].font = Font(name='Calibri', size=10, italic=True)

ct_cols = ['#', 'Name', 'Round', 'P1 CT', 'Parcell CT', 'CT Verified', 'Parcell ID', 'Parcell Address']
for col, h in enumerate(ct_cols, 1):
    c = ws_ct.cell(row=4, column=col, value=h)
    c.font = hf; c.fill = bf; c.border = bdr; c.alignment = ca

for i, r in enumerate(records):
    rw = i + 5
    if r['ct_verified']:
        ct_label = '✅ CT MATCH'
    elif r['ct_mismatch']:
        ct_label = '⚠️ CT MISMATCH'
    else:
        ct_label = '❌ NO RESULTS'
    
    vals = [i+1, r['name'], r['round'], r['p1_ct'], r['p_ct'], ct_label, r['pipe_pid'], r['pipe_addr']]
    for col, v in enumerate(vals, 1):
        c = ws_ct.cell(row=rw, column=col, value=v)
        c.font = nf; c.border = bdr
        if col == 6:
            c.alignment = ca
            if 'MATCH' in str(v): c.fill = gf
            elif 'MISMATCH' in str(v): c.fill = rf
            elif 'NO RESULTS' in str(v): c.fill = gr

ct_widths = [5, 40, 8, 14, 14, 18, 14, 35]
for j, w in enumerate(ct_widths):
    ws_ct.column_dimensions[get_column_letter(j+1)].width = w
ws_ct.auto_filter.ref = f'A4:H{total+4}'
ws_ct.freeze_panes = 'A5'

wb.save(args.out)

# Stats
print(f'📊 {args.out}')
print(f'   Sheets: {wb.sheetnames}')
print()
print('Match rates:')
for label, key, title, desc in match_fields:
    c = Counter(r[key] for r in records)
    yes = c.get('YES',0); no = c.get('NO',0)
    rate = f'{yes/(yes+no)*100:.1f}%' if yes+no else 'N/A'
    print(f'   {title}: YES={yes} NO={no} N/A={c.get("N/A",0)} → {rate}')
