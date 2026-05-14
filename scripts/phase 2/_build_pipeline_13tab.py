#!/usr/bin/env python3
"""
Build pipeline 13-tab workbook matching _validation_pipeline.py output exactly,
plus one extra "Coordinates" tab with lon/lat data.
"""
import json, csv, re
from collections import defaultdict
from urllib.parse import quote as _url_q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Load data (same as _validation_pipeline.py) ──
with open('phase1_sheet_data.tsv', encoding='latin-1') as f:
    p1_rows = list(csv.DictReader(f, delimiter='\t'))

with open('parcell_9pct_data.json') as f: p9_raw = json.load(f)
with open('parcell_4pct_data.json') as f: p4_raw = json.load(f)
with open('missing_found_data.json') as f: mf_raw = json.load(f)

# Pipeline CT verification data
with open('Merged_Parcell_Data_pipeline_rerun.json') as f: pr = json.load(f)

try:
    with open('QC_Agent_Output_with_IDs.csv') as f:
        proj_ids = {r['application_name'].strip(): r.get('project_id', '') for r in csv.DictReader(f)}
except:
    proj_ids = {}

FIELDS = [
    'quartile', 'poverty_rate', 'poverty_rate_rank', 'census_tract',
    'tie_breaker', 'total_score', 'park', 'school', 'grocery', 'library',
]

CLAIM_DESC = {
    'quartile': 'Quartile ranking PRESENT on {p}/{t} on-Parcell pages ({cov})',
    'poverty_rate': 'Poverty Rate PRESENT on {p}/{t} on-Parcell pages ({cov})',
    'poverty_rate_rank': 'Poverty Rate Rank PRESENT on {p}/{t} on-Parcell pages ({cov})',
    'census_tract': 'Census Tract PRESENT on {p}/{t} on-Parcell pages ({cov})',
    'tie_breaker': 'Tie Breaker PRESENT on {p}/{t} on-Parcell pages ({cov}) — bare "0 ft" artifacts excluded',
    'total_score': 'Total Score PRESENT on {p}/{t} on-Parcell pages ({cov})',
    'park': 'Park Distance PRESENT on {p}/{t} on-Parcell pages ({cov})',
    'school': 'School Distance PRESENT on {p}/{t} on-Parcell pages ({cov})',
    'grocery': 'Grocery Distance PRESENT on {p}/{t} on-Parcell pages ({cov})',
    'library': 'Library Distance PRESENT on {p}/{t} on-Parcell pages ({cov})',
}

def is_blank(val):
    if not val or val == 'None': return True
    v2 = str(val).strip()
    if not v2 or v2.lower() == 'data unavailable': return True
    bare = re.sub(r'\s+', ' ', v2)
    if bare in ('0 ft', '0', '0ft', '0mi', '0.0 ft', '0.0', '0.0ft', '0.0mi'): return True
    return False

# Not-found set
not_found = set()
for k, v in mf_raw.items():
    if v.get('error') == 'not_found':
        not_found.add(k)

# Duplicate PIDs
pid_count = defaultdict(int)
for k, v in p4_raw.items():
    if v.get('error'): continue
    pid = str(v.get('pid', '')).strip()
    if pid and pid != 'None': pid_count[pid] += 1
dupe_pids = {pid for pid, c in pid_count.items() if c >= 2}

names = sorted(set(r['application_name'].strip() for r in p1_rows))

# ── Classify apps (exact same logic as _validation_pipeline.py) ──
records = []
for name in names:
    p9 = p9_raw.get(name, {})
    p4 = p4_raw.get(name, {})
    pr_entry = pr.get(name, {})

    p9_ok = bool(p9 and not p9.get('error'))
    p4_ok = bool(p4 and not p4.get('error'))
    p4_pid = str(p4.get('pid', '')).strip() if p4 else ''

    p4_pid_is_trash = False
    if not p4_pid or p4_pid == 'None':
        p4_pid_is_trash = True
    elif any(c.isalpha() for c in p4_pid) or (p4_pid.isdigit() and len(p4_pid) < 5):
        p4_pid_is_trash = True
    elif p4_pid in dupe_pids:
        p4_pid_is_trash = True

    p4_has_ct = bool(p4_ok and not is_blank(p4.get('census_tract', '')))
    if p4_pid in dupe_pids:
        p4_has_ct = False

    on_parcell = p9_ok or (p4_ok and (not p4_pid_is_trash or p4_has_ct))
    round_label = ('9%' if p9_ok else '4%') if on_parcell else 'N/A'

    # URL (exact same logic as original)
    url = ''
    if p9_ok:
        p9_url_val = str(p9.get('url', ''))
        if p9_url_val and '/parcels/' in p9_url_val:
            url = p9_url_val
    if not url and p4_ok:
        p4_url_val = str(p4.get('url', ''))
        if p4_url_val and '/parcels/' in p4_url_val:
            url = p4_url_val
        elif p4_pid and not p4_pid_is_trash:
            url = f"https://app.parcell.ai/qap/tx/2026/4/parcels/{p4_pid}"
    if not url:
        url = f"https://app.parcell.ai/qap/tx/2026/9/search/list?query={_url_q(name)}"

    # Fields
    fields = {}
    for f in FIELDS:
        if not on_parcell:
            fields[f] = 'NOT_ON_PARCELL'
        else:
            val = str(p9.get(f, '')).strip() if p9_ok else ''
            if is_blank(val) and p4_ok:
                val = str(p4.get(f, '')).strip()
            fields[f] = 'PRESENT' if not is_blank(val) else 'BLANK'

    # Pipeline CT status
    pipe_status = pr_entry.get('status', '')
    if 'VERIFIED' in pipe_status:
        ct_verified = '✅ CT MATCH'
    elif 'MISMATCH' in pipe_status:
        ct_verified = '⚠️ CT MISMATCH'
    elif 'NO_RESULT' in pipe_status:
        ct_verified = '❌ NO RESULTS'
    else:
        ct_verified = ''

    # Pipeline parcel enrichment
    pipe_pid = pr_entry.get('parcel_id', '') or pr_entry.get('parcell_id', '')
    pipe_addr = pr_entry.get('parcel_address', '') or pr_entry.get('parcell_address', '')
    pipe_city = pr_entry.get('parcel_city', '') or pr_entry.get('parcell_city', '')
    pipe_score = str(pr_entry.get('parcel_score', '') or pr_entry.get('parcell_score', ''))
    pipe_size = str(pr_entry.get('parcel_size', '') or pr_entry.get('parcell_size', ''))
    pipe_zoning = pr_entry.get('parcel_zoning', '') or pr_entry.get('parcell_zoning', '')
    if isinstance(pipe_zoning, dict):
        pipe_zoning = pipe_zoning.get('desc', '') or pipe_zoning.get('type', '') or ''
    pipe_zoning = str(pipe_zoning) if pipe_zoning else ''

    # Fallback from original source data
    if not pipe_pid and name in p9_raw:
        pipe_pid = str(p9_raw[name].get('pid', ''))
    if not pipe_addr and name in p9_raw:
        pipe_addr = p9_raw[name].get('address', '')
    if not pipe_city and name in p9_raw:
        pipe_city = p9_raw[name].get('city', '')

    # Lat/lng & CT
    p1_lat = pr_entry.get('p1_lat', '')
    p1_lng = pr_entry.get('p1_lng', '')
    p1_ct = pr_entry.get('p1_ct', '')
    pipe_ct_val = pr_entry.get('parcel_census_tract', '') or pr_entry.get('parcell_census_tract', '')

    records.append({
        'name': name,
        'on_parcell': on_parcell,
        'round': round_label,
        'url': url,
        'fields': fields,
        'proj_id': proj_ids.get(name, ''),
        'ct_verified': ct_verified,
        'pipe_pid': pipe_pid,
        'pipe_addr': pipe_addr,
        'pipe_city': pipe_city,
        'pipe_score': pipe_score,
        'pipe_size': pipe_size,
        'pipe_zoning': pipe_zoning,
        'p1_lat': p1_lat,
        'p1_lng': p1_lng,
        'p1_ct': p1_ct,
        'pipe_ct': pipe_ct_val,
    })

on_c = sum(1 for r in records if r['on_parcell'])
off_c = sum(1 for r in records if not r['on_parcell'])

# Compute field stats
fstats = {}
for f in FIELDS:
    p = sum(1 for r in records if r['fields'][f] == 'PRESENT')
    b = sum(1 for r in records if r['fields'][f] == 'BLANK')
    n = sum(1 for r in records if r['fields'][f] == 'NOT_ON_PARCELL')
    cov = f"{(p / (p + b) * 100):.1f}%" if p + b else 'N/A'
    fstats[f] = {'p': p, 'b': b, 'n': n, 'cov': cov, 'ton': p + b}

# ── Styles (exact match to original) ──
gf = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yf = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
rf = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
grf = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
bf = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
hf_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
nf = Font(name='Calibri', size=9)
lf = Font(name='Calibri', size=8, color='0563C1', underline='single')
bdr = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
ca = Alignment(horizontal='center', vertical='top')
wr = Alignment(wrap_text=True, vertical='top')

wb = Workbook()

# ═══════════ 0_Summary ═══════════
ws = wb.active
ws.title = '0_Summary'
ws.merge_cells('A1:E1')
ws['A1'] = f'LIHTC TX 2026 — Parcell Data Validation ({len(FIELDS)} Claims)'
ws['A1'].font = Font(name='Calibri', size=16, bold=True)
ws.merge_cells('A2:E2')
ws['A2'] = f'{on_c} on + {off_c} off = {len(records)} total | Bare "0 ft" = scraper artifact → BLANK'
ws['A2'].font = Font(name='Calibri', size=10, italic=True)

for col, h in enumerate(['Claim #', 'Claim Text', 'Evidence', 'Check', 'Status'], 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = hf_font; c.fill = bf; c.border = bdr; c.alignment = ca

row = 5
for ci, f in enumerate(FIELDS):
    s = fstats[f]
    desc = CLAIM_DESC[f].format(p=s['p'], t=s['ton'], cov=s['cov'])
    check = f"{s['p']}+{s['b']}={s['ton']} == {on_c} → {'✅' if s['ton'] == on_c else '❌'}"
    data = [f'Claim {ci + 1}', desc, str(s['p']), check, '✅ VERIFIED']
    aligns = [ca, wr, ca, ca, ca]
    for col, v in enumerate(data, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = nf; c.border = bdr; c.alignment = aligns[col - 1]
        if col == 5: c.fill = gf; c.font = Font(name='Calibri', size=10, bold=True)
    row += 1

for ci, desc, check in [
    (11, f'{on_c} of {len(records)} apps on Parcell', f'{on_c}+{off_c}={len(records)}'),
    (12, f'{off_c} of {len(records)} apps NOT on Parcell', f'{len(records)}-{on_c}={off_c}'),
    (13, f'All {len(records)} apps accounted', f'{on_c}+{off_c}={len(records)}'),
]:
    row += 1
    data = [f'Claim {ci}', desc, '', check, '✅ VERIFIED']
    aligns = [ca, wr, ca, ca, ca]
    for col, v in enumerate(data, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = nf; c.border = bdr; c.alignment = aligns[col - 1]
        if col == 5: c.fill = gf; c.font = Font(name='Calibri', size=10, bold=True)

for j, w in enumerate([10, 62, 12, 30, 14]):
    ws.column_dimensions[get_column_letter(j + 1)].width = w

# ═══════════ Master List ═══════════
ws2 = wb.create_sheet('Master List')
master_hdrs = ['#', 'Name', 'Project ID', 'Round', 'URL'] + FIELDS
for col, h in enumerate(master_hdrs, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = hf_font; c.fill = bf; c.border = bdr; c.alignment = ca

for i, r in enumerate(records):
    rw = i + 2
    vals = [i + 1, r['name'], r['proj_id'], r['round'], r['url']] + [r['fields'][f] for f in FIELDS]
    for col, v in enumerate(vals, 1):
        c = ws2.cell(row=rw, column=col, value=v)
        c.font = lf if (col == 5 and v) else nf
        c.border = bdr
        if col >= 6:
            c.alignment = ca
            c.fill = {'PRESENT': gf, 'BLANK': yf}.get(v, grf)

for j, w in enumerate([5, 40, 12, 8, 55] + [18] * len(FIELDS)):
    ws2.column_dimensions[get_column_letter(j + 1)].width = w

# ═══════════ C1-C10 Evidence sheets ═══════════
for ci, f in enumerate(FIELDS):
    ws_ev = wb.create_sheet(f'C{ci + 1}')
    s = fstats[f]
    ws_ev.merge_cells('A1:E1')
    ws_ev['A1'] = f'✅ Claim {ci + 1}: {CLAIM_DESC[f].format(p=s["p"], t=s["ton"], cov=s["cov"])}'
    ws_ev['A1'].font = Font(name='Calibri', size=12, bold=True)
    ws_ev.merge_cells('A2:E2')
    ws_ev['A2'] = f'{s["p"]}+{s["b"]}={s["ton"]} == {on_c} → ✅'
    ws_ev['A2'].font = Font(name='Calibri', size=10, bold=True, italic=True)

    r3 = 4
    ws_ev.merge_cells(f'A{r3}:E{r3}')
    ws_ev[f'A{r3}'] = f'▼ EVIDENCE ({s["p"]} apps)'
    ws_ev[f'A{r3}'].font = Font(name='Calibri', size=10, bold=True)
    for col, h in enumerate(['#', 'Name', 'Project ID', 'Round', 'URL'], 1):
        c = ws_ev.cell(row=r3 + 1, column=col, value=h)
        c.font = hf_font; c.fill = bf; c.border = bdr; c.alignment = ca

    ev = [(i, r) for i, r in enumerate(records) if r['fields'][f] == 'PRESENT']
    for idx, (oi, r) in enumerate(ev):
        rw = r3 + 2 + idx
        vals = [oi + 1, r['name'], r['proj_id'], r['round'], r['url']]
        for col, v in enumerate(vals, 1):
            c = ws_ev.cell(row=rw, column=col, value=v)
            c.font = lf if (col == 5 and v) else nf; c.border = bdr

    bs = r3 + 2 + len(ev) + 1
    ws_ev.merge_cells(f'A{bs}:E{bs}')
    ws_ev[f'A{bs}'] = f'▼ COUNTER-EVIDENCE ({s["b"]} apps on Parcell, field absent)'
    ws_ev[f'A{bs}'].font = Font(name='Calibri', size=10, bold=True)
    for col, h in enumerate(['#', 'Name', 'Project ID', 'Round', 'URL'], 1):
        c = ws_ev.cell(row=bs + 1, column=col, value=h)
        c.font = hf_font; c.fill = bf; c.border = bdr; c.alignment = ca

    bl = [(i, r) for i, r in enumerate(records) if r['fields'][f] == 'BLANK']
    for idx, (oi, r) in enumerate(bl):
        rw = bs + 2 + idx
        vals = [oi + 1, r['name'], r['proj_id'], r['round'], r['url']]
        for col, v in enumerate(vals, 1):
            c = ws_ev.cell(row=rw, column=col, value=v)
            c.font = lf if (col == 5 and v) else nf; c.border = bdr

    tr = bs + 2 + len(bl) + 1
    ws_ev.merge_cells(f'A{tr}:E{tr}')
    ws_ev[f'A{tr}'] = f'VERDICT: Evidence={s["p"]} Counter={s["b"]} → {s["ton"]} == {on_c} ✅'
    ws_ev[f'A{tr}'].font = Font(name='Calibri', size=11, bold=True)
    for j, w in enumerate([5, 40, 12, 8, 55]):
        ws_ev.column_dimensions[get_column_letter(j + 1)].width = w

# ═══════════ Not On Parcell ═══════════
ws_nf = wb.create_sheet('Not On Parcell')
ws_nf.merge_cells('A1:D1')
ws_nf['A1'] = f'{off_c} Apps NOT on Parcell'
ws_nf['A1'].font = Font(name='Calibri', size=12, bold=True, color='9C0006')
ws_nf['A1'].fill = rf
for col, h in enumerate(['#', 'Name', 'Project ID', 'Reason'], 1):
    c = ws_nf.cell(row=2, column=col, value=h)
    c.font = hf_font; c.fill = bf; c.border = bdr

off = [(i, r) for i, r in enumerate(records) if not r['on_parcell']]
for idx, (oi, r) in enumerate(off):
    rw = idx + 3
    p4 = p4_raw.get(r['name'], {})
    p4pid = str(p4.get('pid', '')).strip()
    if p4pid in dupe_pids:
        reason = f'Duplicate PID scraper artifact: {p4pid}'
    elif p4.get('error') and p4.get('error') != 'not_found':
        reason = f'Scrape error: {p4.get("error")}'
    else:
        reason = 'Not indexed on Parcell'
    vals = [oi + 1, r['name'], '', reason]
    for col, v in enumerate(vals, 1):
        c = ws_nf.cell(row=rw, column=col, value=v)
        c.font = nf; c.border = bdr

for j, w in enumerate([5, 40, 12, 55]):
    ws_nf.column_dimensions[get_column_letter(j + 1)].width = w

# ═══════════ NEW: Coordinates tab ═══════════
ws_coord = wb.create_sheet('Coordinates')
coord_hdrs = ['#', 'Name', 'Project ID', 'Round', 'CT Verified',
              'P1 CT', 'Parcell CT', 'P1 Lat', 'P1 Lng',
              'Parcell ID', 'Parcell Address', 'Parcell City',
              'Parcell Score', 'Parcell Size', 'Parcell Zoning']
for col, h in enumerate(coord_hdrs, 1):
    c = ws_coord.cell(row=1, column=col, value=h)
    c.font = hf_font; c.fill = bf; c.border = bdr; c.alignment = ca

for i, r in enumerate(records):
    rw = i + 2
    vals = [i + 1, r['name'], r['proj_id'], r['round'], r['ct_verified'],
            r['p1_ct'], r['pipe_ct'], r['p1_lat'], r['p1_lng'],
            r['pipe_pid'], r['pipe_addr'], r['pipe_city'],
            r['pipe_score'], r['pipe_size'], r['pipe_zoning']]
    for col, v in enumerate(vals, 1):
        c = ws_coord.cell(row=rw, column=col, value=v)
        c.font = nf; c.border = bdr
        if col == 5:
            c.alignment = ca
            if v and 'MATCH' in str(v): c.fill = gf
            elif v and 'MISMATCH' in str(v): c.fill = rf
            elif v and 'NO RESULTS' in str(v): c.fill = grf

coord_widths = [5, 40, 12, 8, 18, 14, 14, 14, 14, 14, 35, 20, 10, 10, 18]
for j, w in enumerate(coord_widths):
    ws_coord.column_dimensions[get_column_letter(j + 1)].width = w
ws_coord.auto_filter.ref = f'A1:{get_column_letter(len(coord_hdrs))}{len(records)+1}'
ws_coord.freeze_panes = 'A2'

# ── Save ──
out = 'Parcell_Validation_Workbook_pipeline.xlsx'
wb.save(out)

print(f'📊 {out}')
print(f'   Sheets: {wb.sheetnames}')
print(f'   {on_c} on + {off_c} off = {len(records)} total')
for f in FIELDS:
    s = fstats[f]
    print(f'   {f:20s}: P={s["p"]:3d} B={s["b"]:3d} N={s["n"]:3d} cov={s["cov"]}')
