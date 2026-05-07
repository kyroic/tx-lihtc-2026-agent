from __future__ import annotations

import csv
import hashlib
import json
import os
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .model_client import chat_completions, extract_json_content

# Excel per-cell limit (inclusive); stay under for long review_reasons strings.
_EXCEL_MAX_CELL_LEN = 32000


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def norm_ws(s: str) -> str:
    return " ".join((s or "").split())


def coaching_append_from_env() -> str:
    """
    Text from LIHTC_COACHING_APPEND (e.g. set by eval --coaching-file or progression parent).
    Appended to extraction system prompts so later iterations can apply OpenClaw coaching.
    """
    raw = (os.environ.get("LIHTC_COACHING_APPEND") or "").strip()
    if not raw:
        return ""
    return (
        "\n\n=== Iteration coaching (from prior eval; follow strictly) ===\n"
        + norm_ws(raw)[:12000]
    )


def read_pdf_pages(pdf_path: Path, max_pages: int) -> list[dict[str, Any]]:
    pages: list[dict[str, Any]] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for idx, page in enumerate(pdf.pages[:max_pages], start=1):
            try:
                text = page.extract_text() or ""
            except Exception:
                text = ""
            pages.append({"page": idx, "text": (text or "").strip()[:4000]})
    return pages


def build_page_hints(pages: list[dict[str, Any]]) -> dict[str, list[int]]:
    """
    Heuristic locator for likely sections/pages.

    Returns dict[field -> list[page_numbers]].
    """
    def has_any(t: str, needles: list[str]) -> bool:
        return any(n in t for n in needles)

    hints: dict[str, list[int]] = {
        "application_name": [],
        "contact": [],
        "census_tract": [],
        "quartile": [],
        "property_rate": [],
        "poverty_rank": [],
        "tiebreaker_park": [],
        "tiebreaker_school": [],
        "tiebreaker_grocery": [],
        "tiebreaker_library": [],
        "scoring": [],
        "site": [],
    }

    for p in pages:
        pn = int(p.get("page") or 0)
        txt = norm_ws(str(p.get("text") or "")).lower()
        if not pn or not txt:
            continue

        if has_any(txt, ["development name", "project name", "property name", "application name"]):
            hints["application_name"].append(pn)
        if has_any(txt, ["contact", "prepared by", "authorized representative", "applicant contact", "developer contact"]):
            hints["contact"].append(pn)

        if has_any(txt, ["census tract", "tract", "geoid", "geo id", "geographic id", "block group", "fips"]):
            hints["census_tract"].append(pn)

        if has_any(txt, ["quartile", "income quartile", "q1", "q2", "q3", "q4", "qhfd", "qhfa"]):
            hints["quartile"].append(pn)

        if has_any(txt, ["tax rate", "property tax", "property rate", "mill rate", "mills", "taxes per", "levy"]):
            hints["property_rate"].append(pn)

        if has_any(
            txt,
            [
                "poverty",
                "federal poverty",
                "fpl",
                "fpg",
                "% of poverty",
                "percent of poverty",
                "median income",
                "ami",
                "area median",
                "income level",
                "mfi",
                "median family income",
            ],
        ):
            hints["poverty_rank"].append(pn)

        if has_any(txt, ["tie-break", "tiebreak", "tie breaker", "scoring", "score"]):
            hints["scoring"].append(pn)

        if has_any(txt, ["site", "location", "address", "census", "tract"]):
            hints["site"].append(pn)

        if has_any(txt, ["park", "playground"]):
            hints["tiebreaker_park"].append(pn)
        if "school" in txt:
            hints["tiebreaker_school"].append(pn)
        if has_any(txt, ["grocery", "supermarket"]):
            hints["tiebreaker_grocery"].append(pn)
        if "library" in txt:
            hints["tiebreaker_library"].append(pn)

    # De-dup and keep stable order
    for k, v in list(hints.items()):
        seen = set()
        out = []
        for x in v:
            if x not in seen:
                seen.add(x)
                out.append(x)
        hints[k] = out[:10]
    return hints


@dataclass
class FieldEvidence:
    value: str = ""
    confidence: float = 0.0
    pages: list[int] = field(default_factory=list)
    quote: str = ""


@dataclass
class ExtractedRow:
    source_pdf_path: str
    source_pdf_sha256: str
    extraction_version: str = "agent-v0"

    application_name: FieldEvidence = field(default_factory=FieldEvidence)
    contact_name: FieldEvidence = field(default_factory=FieldEvidence)
    contact_email: FieldEvidence = field(default_factory=FieldEvidence)
    contact_phone: FieldEvidence = field(default_factory=FieldEvidence)

    tiebreaker_park: FieldEvidence = field(default_factory=FieldEvidence)
    tiebreaker_school: FieldEvidence = field(default_factory=FieldEvidence)
    tiebreaker_grocery: FieldEvidence = field(default_factory=FieldEvidence)
    tiebreaker_library: FieldEvidence = field(default_factory=FieldEvidence)

    quartile: FieldEvidence = field(default_factory=FieldEvidence)
    property_rate: FieldEvidence = field(default_factory=FieldEvidence)
    poverty_rank: FieldEvidence = field(default_factory=FieldEvidence)
    census_tract: FieldEvidence = field(default_factory=FieldEvidence)

    needs_review: bool = True
    review_reasons: list[str] = field(default_factory=list)


_EVIDENCE_KEYS = (
    "application_name",
    "contact_name",
    "contact_email",
    "contact_phone",
    "tiebreaker_park",
    "tiebreaker_school",
    "tiebreaker_grocery",
    "tiebreaker_library",
    "quartile",
    "property_rate",
    "poverty_rank",
    "census_tract",
)


def extracted_row_from_jsondict(d: dict[str, Any]) -> ExtractedRow:
    """Rebuild ExtractedRow from ``asdict(row)`` / applications.jsonl line."""
    row = ExtractedRow(
        source_pdf_path=str(d.get("source_pdf_path") or ""),
        source_pdf_sha256=str(d.get("source_pdf_sha256") or ""),
        extraction_version=str(d.get("extraction_version") or "agent-v0"),
    )
    for k in _EVIDENCE_KEYS:
        v = d.get(k)
        if isinstance(v, dict):
            setattr(row, k, field_from_obj(v))
        else:
            setattr(row, k, FieldEvidence())
    row.needs_review = bool(d.get("needs_review"))
    rr = d.get("review_reasons")
    if isinstance(rr, list):
        row.review_reasons = [str(x) for x in rr]
    elif isinstance(rr, str) and rr.strip():
        row.review_reasons = [s for s in rr.split(";") if s]
    return row


def field_from_obj(obj: dict[str, Any] | None) -> FieldEvidence:
    obj = obj or {}
    pages = obj.get("pages") or []
    out_pages: list[int] = []
    for p in pages:
        try:
            out_pages.append(int(p))
        except Exception:
            continue
    return FieldEvidence(
        value=str(obj.get("value") or "").strip(),
        confidence=float(obj.get("confidence") or 0.0),
        pages=out_pages,
        quote=norm_ws(str(obj.get("quote") or ""))[:240],
    )


def extract_one_pdf(*, project_id: str, model: str, pdf_path: Path, max_pages: int) -> ExtractedRow:
    pages = read_pdf_pages(pdf_path, max_pages=max_pages)
    page_hints = build_page_hints(pages)

    system = (
        "You are an extraction agent for Texas LIHTC 2026 Full Application PDFs.\n"
        "Return ONLY valid JSON (no markdown).\n"
        "Rules:\n"
        "- Never invent values. If not present, return value=\"\" and confidence=0.\n"
        "- Every non-empty value MUST include pages[] and a short quote copied from the PDF text.\n"
        "- Use page_hints to focus on the likely scoring/site/tract/tie-breaker sections.\n"
    ) + coaching_append_from_env()
    schema_hint = {
        "application_name": {"value": "", "confidence": 0.0, "pages": [], "quote": ""},
        "contact_name": {"value": "", "confidence": 0.0, "pages": [], "quote": ""},
        "contact_email": {"value": "", "confidence": 0.0, "pages": [], "quote": ""},
        "contact_phone": {"value": "", "confidence": 0.0, "pages": [], "quote": ""},
        "tiebreaker_park": {"value": "", "confidence": 0.0, "pages": [], "quote": ""},
        "tiebreaker_school": {"value": "", "confidence": 0.0, "pages": [], "quote": ""},
        "tiebreaker_grocery": {"value": "", "confidence": 0.0, "pages": [], "quote": ""},
        "tiebreaker_library": {"value": "", "confidence": 0.0, "pages": [], "quote": ""},
        "quartile": {"value": "", "confidence": 0.0, "pages": [], "quote": ""},
        "property_rate": {"value": "", "confidence": 0.0, "pages": [], "quote": ""},
        "poverty_rank": {"value": "", "confidence": 0.0, "pages": [], "quote": ""},
        "census_tract": {"value": "", "confidence": 0.0, "pages": [], "quote": ""},
    }
    user = json.dumps(
        {
            "pdf_filename": pdf_path.name,
            "page_hints": page_hints,
            "pages": pages,
            "output_schema_example": schema_hint,
        },
        ensure_ascii=False,
    )

    resp = chat_completions(
        project_id=project_id,
        model=model,
        messages=[{"role": "system", "content": system}, {"role": "user", "content": user}],
        temperature=0.0,
    )
    out = extract_json_content(resp)

    row = ExtractedRow(source_pdf_path=str(pdf_path), source_pdf_sha256=sha256_file(pdf_path))
    for k in schema_hint.keys():
        setattr(row, k, field_from_obj(out.get(k)))

    required = {
        "application_name": row.application_name.value,
        "contact_email": row.contact_email.value,
        "census_tract": row.census_tract.value,
    }
    for k, v in required.items():
        if not (v or "").strip():
            row.review_reasons.append(f"missing:{k}")
    for k in schema_hint.keys():
        f = getattr(row, k)
        if f.value and (not f.pages or not f.quote.strip()):
            row.review_reasons.append(f"missing_evidence:{k}")
    row.needs_review = bool(row.review_reasons)
    return row


def csv_fieldnames() -> list[str]:
    return [
        "source_pdf_path",
        "source_pdf_sha256",
        "extraction_version",
        "application_name",
        "contact_name",
        "contact_email",
        "contact_phone",
        "tiebreaker_park",
        "tiebreaker_school",
        "tiebreaker_grocery",
        "tiebreaker_library",
        "quartile",
        "property_rate",
        "poverty_rank",
        "census_tract",
        "needs_review",
        "review_reasons",
    ]


def excel_fieldnames() -> list[str]:
    """One row per application PDF: CLI project scope + all extracted columns."""
    return ["project_id", *csv_fieldnames()]


def flatten_excel_row(row: ExtractedRow, *, project_id: str) -> dict[str, str]:
    base = flatten_csv(row)
    return {"project_id": project_id, **base}


def _truncate_cell(s: str) -> str:
    if len(s) <= _EXCEL_MAX_CELL_LEN:
        return s
    return s[: _EXCEL_MAX_CELL_LEN - 3] + "..."


def write_applications_xlsx(*, path: Path, project_id: str, rows: list[ExtractedRow]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = excel_fieldnames()
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(fieldnames)
    for r in rows:
        flat = flatten_excel_row(r, project_id=project_id)
        ws.append([_truncate_cell(str(flat.get(c, "") or "")) for c in fieldnames])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{len(rows) + 1}"
    wb.save(str(path))


def flatten_csv(row: ExtractedRow) -> dict[str, str]:
    def v(x: FieldEvidence) -> str:
        return x.value

    return {
        "source_pdf_path": row.source_pdf_path,
        "source_pdf_sha256": row.source_pdf_sha256,
        "extraction_version": row.extraction_version,
        "application_name": v(row.application_name),
        "contact_name": v(row.contact_name),
        "contact_email": v(row.contact_email),
        "contact_phone": v(row.contact_phone),
        "tiebreaker_park": v(row.tiebreaker_park),
        "tiebreaker_school": v(row.tiebreaker_school),
        "tiebreaker_grocery": v(row.tiebreaker_grocery),
        "tiebreaker_library": v(row.tiebreaker_library),
        "quartile": v(row.quartile),
        "property_rate": v(row.property_rate),
        "poverty_rank": v(row.poverty_rank),
        "census_tract": v(row.census_tract),
        "needs_review": "true" if row.needs_review else "false",
        "review_reasons": ";".join(row.review_reasons),
    }


def write_outputs(*, out_dir: Path, rows: list[ExtractedRow], project_id: str, model: str, max_pages: int) -> dict[str, Any]:
    out_dir.mkdir(parents=True, exist_ok=True)
    apps_csv = out_dir / "applications.csv"
    apps_xlsx = out_dir / "applications.xlsx"
    apps_jsonl = out_dir / "applications.jsonl"
    review_csv = out_dir / "review_queue.csv"
    run_summary = out_dir / "run_summary.json"

    with apps_jsonl.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(asdict(r), ensure_ascii=False) + "\n")

    fieldnames = csv_fieldnames()
    flat = [flatten_csv(r) for r in rows]
    with apps_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(flat)

    write_applications_xlsx(path=apps_xlsx, project_id=project_id, rows=rows)

    review = [x for x in flat if x["needs_review"] == "true"]
    with review_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(review)

    summary = {
        "project_id": project_id,
        "count_pdfs": len(rows),
        "count_needs_review": len(review),
        "model": model,
        "max_pages": max_pages,
        "outputs": {
            "applications_xlsx": str(apps_xlsx),
            "applications_csv": str(apps_csv),
            "applications_jsonl": str(apps_jsonl),
            "review_queue_csv": str(review_csv),
        },
    }
    run_summary.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    return summary

